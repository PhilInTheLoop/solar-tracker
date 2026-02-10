"""
Script to import historical data from Excel file
"""
from openpyxl import load_workbook
from datetime import datetime
from pathlib import Path
import sqlite3

EXCEL_PATH = Path(__file__).parent.parent / "Sonnenertrag.xlsx"
DB_PATH = Path(__file__).parent / "backend" / "solar_data.db"

def import_excel_data():
    print(f"Reading Excel from: {EXCEL_PATH}")

    # Read Excel
    wb = load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['Tabelle1']

    # Connect to database
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    # Create tables if not exist
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS readings (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL UNIQUE,
            meter_reading REAL NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    # Extract settings from Excel
    settings = {
        'plant_size_kwp': '4.84',
        'price_per_kwh': '0.518',
        'expected_yield_per_kwp': '950',
        'start_date': '2006-04-20',
        'initial_meter_reading': '2110.5',
        'address': 'Deutschland',
        'latitude': '48.1351',
        'longitude': '11.5820',
        'currency': 'EUR'
    }

    for key, value in settings.items():
        cursor.execute('''
            INSERT OR REPLACE INTO settings (key, value, updated_at)
            VALUES (?, ?, CURRENT_TIMESTAMP)
        ''', (key, value))

    # Extract readings (rows starting from row 7, index 0-based is 6)
    readings_count = 0
    for row in ws.iter_rows(min_row=7, values_only=True):
        try:
            date_val = row[0]
            meter_val = row[1]

            if date_val is None or meter_val is None:
                continue

            # Check if it's a date
            if isinstance(date_val, datetime):
                date_str = date_val.strftime('%Y-%m-%d')
                if isinstance(meter_val, (int, float)) and meter_val > 0:
                    cursor.execute('''
                        INSERT OR REPLACE INTO readings (date, meter_reading)
                        VALUES (?, ?)
                    ''', (date_str, float(meter_val)))
                    readings_count += 1
        except Exception as e:
            continue

    conn.commit()
    conn.close()

    print(f"Imported {readings_count} readings")
    print("Settings configured:")
    for k, v in settings.items():
        print(f"  {k}: {v}")

if __name__ == "__main__":
    import_excel_data()

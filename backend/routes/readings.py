from fastapi import APIRouter, HTTPException, UploadFile, File
from pydantic import BaseModel
from typing import Optional
from datetime import datetime
from openpyxl import load_workbook
import io

from ..database import (
    get_all_readings, add_reading, delete_reading,
    get_all_settings, import_readings_bulk
)

router = APIRouter(prefix="/api/readings", tags=["readings"])


def calculate_yield(current_reading, prev_reading, date, prev_date, meter_change_date):
    """Calculate yield handling meter changes"""
    # Check if this is the first reading after meter change
    if meter_change_date and date >= meter_change_date and (not prev_date or prev_date < meter_change_date):
        # First reading after meter change - yield is just the new meter value
        return current_reading
    # Check if meter was reset (reading drops significantly)
    elif prev_reading and current_reading < prev_reading and prev_reading - current_reading > 1000:
        return current_reading
    else:
        return current_reading - (prev_reading or 0)

class ReadingCreate(BaseModel):
    date: str  # Format: YYYY-MM-DD
    meter_reading: float

class ReadingResponse(BaseModel):
    id: int
    date: str
    meter_reading: float
    yield_kwh: Optional[float] = None
    yield_per_kwp: Optional[float] = None
    revenue: Optional[float] = None

@router.get("")
async def list_readings():
    readings = get_all_readings()
    settings = get_all_settings()

    plant_size = float(settings.get('plant_size_kwp', 4.84))
    price_per_kwh = float(settings.get('price_per_kwh', 0.518))
    initial_reading = float(settings.get('initial_meter_reading', 0))
    meter_change_date = settings.get('meter_change_date', '')

    # Calculate yields
    enriched = []
    prev_reading = initial_reading
    prev_date = ''

    for r in readings:
        current_reading = r['meter_reading']
        yield_kwh = calculate_yield(current_reading, prev_reading, r['date'], prev_date, meter_change_date)

        enriched.append({
            'id': r['id'],
            'date': r['date'],
            'meter_reading': r['meter_reading'],
            'yield_kwh': round(max(0, yield_kwh), 2),
            'yield_per_kwp': round(max(0, yield_kwh) / plant_size, 2) if plant_size > 0 else 0,
            'revenue': round(max(0, yield_kwh) * price_per_kwh, 2)
        })
        prev_reading = current_reading
        prev_date = r['date']

    return enriched

@router.post("")
async def create_reading(reading: ReadingCreate):
    try:
        reading_id = add_reading(reading.date, reading.meter_reading)
        return {"id": reading_id, "message": "Reading added successfully"}
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@router.delete("/{reading_id}")
async def remove_reading(reading_id: int):
    delete_reading(reading_id)
    return {"message": "Reading deleted"}

@router.post("/import-excel")
async def import_from_excel(file: UploadFile = File(...)):
    """Import readings from uploaded Excel file"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="Only Excel files allowed")

    content = await file.read()
    wb = load_workbook(io.BytesIO(content), data_only=True)
    ws = wb.active

    # Find the data rows (looking for dates)
    readings = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        try:
            date_val = row[0]
            meter_val = row[1]

            if date_val is None or meter_val is None:
                continue

            # Check if it's a date
            if isinstance(date_val, datetime):
                date_str = date_val.strftime('%Y-%m-%d')
                if isinstance(meter_val, (int, float)) and meter_val > 0:
                    readings.append({
                        'date': date_str,
                        'meter_reading': float(meter_val)
                    })
        except:
            continue

    if readings:
        import_readings_bulk(readings)
        return {"imported": len(readings), "message": f"Successfully imported {len(readings)} readings"}

    raise HTTPException(status_code=400, detail="No valid readings found in file")

@router.get("/statistics")
async def get_statistics():
    """Get aggregated statistics"""
    readings = get_all_readings()
    settings = get_all_settings()

    if not readings:
        return {
            "total_yield": 0,
            "total_revenue": 0,
            "avg_monthly_yield": 0,
            "years_active": 0,
            "yearly_stats": []
        }

    plant_size = float(settings.get('plant_size_kwp', 4.84))
    price_per_kwh = float(settings.get('price_per_kwh', 0.518))
    initial_reading = float(settings.get('initial_meter_reading', 0))
    expected_yield_per_kwp = float(settings.get('expected_yield_per_kwp', 950))
    meter_change_date = settings.get('meter_change_date', '')
    meter_change_offset = float(settings.get('meter_change_offset', 0))

    # Calculate total yield accounting for meter change
    last_reading = readings[-1]['meter_reading']
    if meter_change_date and readings[-1]['date'] >= meter_change_date:
        # After meter change: offset + current reading - initial
        total_yield = meter_change_offset + last_reading - initial_reading
    else:
        total_yield = last_reading - initial_reading

    total_revenue = total_yield * price_per_kwh

    # Yearly statistics with meter change handling
    yearly_stats = {}
    prev_reading = initial_reading
    prev_date = ''

    for r in readings:
        year = r['date'][:4]
        current_reading = r['meter_reading']

        yield_kwh = calculate_yield(current_reading, prev_reading, r['date'], prev_date, meter_change_date)

        if year not in yearly_stats:
            yearly_stats[year] = {
                'year': int(year),
                'yield_kwh': 0,
                'months': 0,
                'expected_yield': expected_yield_per_kwp * plant_size
            }

        yearly_stats[year]['yield_kwh'] += max(0, yield_kwh)
        yearly_stats[year]['months'] += 1
        prev_reading = current_reading
        prev_date = r['date']

    # Calculate performance for each year
    for year, stats in yearly_stats.items():
        stats['yield_kwh'] = round(stats['yield_kwh'], 2)
        stats['yield_per_kwp'] = round(stats['yield_kwh'] / plant_size, 2)
        stats['revenue'] = round(stats['yield_kwh'] * price_per_kwh, 2)
        stats['performance_pct'] = round((stats['yield_kwh'] / stats['expected_yield']) * 100, 1)

    yearly_list = sorted(yearly_stats.values(), key=lambda x: x['year'])

    # Calculate years active
    if readings:
        first_year = int(readings[0]['date'][:4])
        last_year = int(readings[-1]['date'][:4])
        years_active = last_year - first_year + 1
    else:
        years_active = 0

    return {
        "total_yield": round(total_yield, 2),
        "total_yield_per_kwp": round(total_yield / plant_size, 2) if plant_size > 0 else 0,
        "total_revenue": round(total_revenue, 2),
        "avg_monthly_yield": round(total_yield / len(readings), 2) if readings else 0,
        "years_active": years_active,
        "expected_yearly_yield": round(expected_yield_per_kwp * plant_size, 2),
        "yearly_stats": yearly_list
    }

@router.get("/monthly-comparison")
async def get_monthly_comparison():
    """Get monthly comparison data for charts"""
    readings = get_all_readings()
    settings = get_all_settings()

    initial_reading = float(settings.get('initial_meter_reading', 0))
    meter_change_date = settings.get('meter_change_date', '')

    # Organize by month
    monthly_data = {}
    prev_reading = initial_reading
    prev_date = ''

    for r in readings:
        date = r['date']
        year = int(date[:4])
        month = int(date[5:7])
        current_reading = r['meter_reading']

        yield_kwh = calculate_yield(current_reading, prev_reading, date, prev_date, meter_change_date)

        if month not in monthly_data:
            monthly_data[month] = {'month': month, 'years': {}}

        monthly_data[month]['years'][year] = round(max(0, yield_kwh), 2)
        prev_reading = current_reading
        prev_date = date

    return sorted(monthly_data.values(), key=lambda x: x['month'])
